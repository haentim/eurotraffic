"""Berlin adapter (tier-1, measured hourly).

Monthly ``mq_hr_*.csv.gz`` files hold per-measurement-cross-section hourly counts
(``mq_name``, ``tag`` date, ``stunde`` 0-23, ``q_kfz_mq_hr`` motor-vehicle count).
The ``Stammdaten_*.xlsx`` master file maps each ``MQ_KURZNAME`` to WGS84
coordinates. Density per cross-section/hour = mean ``q_kfz`` over all days.
"""

from __future__ import annotations

import polars as pl

from ..frames import hourly_from_measured
from ..geometry import point_from_xy
from ..registry import CityContext, register

_STAMM_GLOB = "Stammdaten_*.xlsx"
_LON_COL = "LÄNGE (WGS84)"
_LAT_COL = "BREITE (WGS84)"


def _load_coords(ctx: CityContext) -> pl.DataFrame:
    import openpyxl

    path = next(ctx.raw_dir.glob(_STAMM_GLOB))
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    i_name = header.index("MQ_KURZNAME")
    i_lon = header.index(_LON_COL)
    i_lat = header.index(_LAT_COL)

    seen: dict[str, tuple[float, float]] = {}
    for r in rows:
        name = r[i_name]
        lon, lat = r[i_lon], r[i_lat]
        if name and lon is not None and lat is not None and name not in seen:
            seen[name] = (float(lon), float(lat))

    recs = []
    for name, (lon, lat) in seen.items():
        x, y, wkt = point_from_xy(lon, lat)
        recs.append({"mq_name": name, "longitude": x, "latitude": y, "geometry": wkt})
    return pl.DataFrame(recs)


@register("Berlin", tier="measured", note="hourly mq counts + Stammdaten coordinates")
def load(ctx: CityContext) -> pl.DataFrame:
    files = sorted(ctx.raw_dir.glob("mq_hr_*.csv.gz"))
    if not files:
        raise FileNotFoundError("no mq_hr_*.csv.gz files for Berlin")

    parts = []
    for f in files:
        part = pl.read_csv(
            f, separator=";", columns=["mq_name", "stunde", "q_kfz_mq_hr"],
            schema_overrides={"stunde": pl.Int64, "q_kfz_mq_hr": pl.Float64},
        )
        parts.append(part)
    counts = pl.concat(parts, how="vertical")

    density = (
        counts.filter(
            pl.col("stunde").is_between(0, 23) & pl.col("q_kfz_mq_hr").is_not_null()
        )
        .group_by(["mq_name", "stunde"])
        .agg(pl.col("q_kfz_mq_hr").mean().alias("density"))
        .rename({"stunde": "hour"})
    )

    coords = _load_coords(ctx)
    long = (
        density.join(coords, on="mq_name", how="inner")
        .with_columns(
            pl.lit(ctx.city).alias("city"),
            pl.lit(ctx.country).alias("country"),
            pl.col("mq_name").cast(pl.Utf8).alias("sensor_id"),
        )
        .drop("mq_name")
    )
    return hourly_from_measured(long)
